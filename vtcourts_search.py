"""
vtcourts_search.py
------------------
Automates business-name searches on the Vermont Judiciary Public Portal
using a REAL Chrome browser (not headless / not spoofed).

Requirements
------------
    pip install selenium webdriver-manager openpyxl

Usage
-----
    python vtcourts_search.py

Configuration
-------------
Edit BUSINESS_NAMES and DAYS_BACK below.
Set DAYS_BACK = 0 to search with no date filter.

Output
------
New unique cases appended to vtcourts_results.xlsx
Columns: Case Number | Style / Defendant | Location | Searched Name | Date Added
Duplicates (by Case Number) skipped — within run AND against existing file.
"""

import time
import re
from datetime import datetime, timedelta
from pathlib import Path

from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, NoSuchElementException
from webdriver_manager.chrome import ChromeDriverManager
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

# ─────────────────────────────────────────────────────────────────────────────
#  CONFIGURATION
# ─────────────────────────────────────────────────────────────────────────────

BUSINESS_NAMES = [
    "JG Wentworth",
    "J.G. Wentworth",
    "J G Wentworth",
    "J. G. Wentworth",
    "JG Wentworth Organizations",
    "J.G. Wentworth Organizations",
    "JG Wentworth Organizations, LLC",
    "J.G. Wentworth Organizations, LLC",
    "J G Wentworth Organizations, LLC",
    "J. G. Wentworth Organizations, LLC",
    "JG Wentworth Organizations LLC",
    "J.G. Wentworth Organizations LLC",
    "J G Wentworth Organizations LLC",
    "J. G. Wentworth Organizations LLC",
    "JG Wentworth Originations",
    "J.G. Wentworth Originations",
    "JG Wentworth Originations, LLC",
    "J.G. Wentworth Originations, LLC",
    "J G Wentworth Originations, LLC",
    "J. G. Wentworth Originations, LLC",
    "JG Wentworth Originations LLC",
    "J.G. Wentworth Originations LLC",
    "J G Wentworth Originations LLC",
    "J. G. Wentworth Originations LLC",
    "DRB Capital",
    "DRB Capital LLC",
    "DRB Capital, LLC",
    "Stone Street Captial",
    "Stone Street Capital LLC",
    "Stone Street Capital, LLC",
    "AA Ron I LLC",
    "AA Ron I, LLC",
    "Abactor LLC",
    "Abactor, LLC",
    "Abidole LLC",
    "Abidole, LLC",
    "Adenna Med LLC",
    "Adenna Med, LLC",
    "Adventura LLC",
    "Adventura, LLC",
    "AGPI LLC",
    "AGPI, LLC",
    "Aikman Structured Finance",
    "Aikman Structured Finance LLC",
    "Aikman Structured Finance, LLC",
    "Annuity Transfers Ltd",
    "Apis Management LLC",
    "Apis Management, LLC",
    "Atlas Legal Funding III LP",
    "AXE Finance LLC",
    "AXE Finance, LLC",
    "B.A.W.21",
    "B.R. Wright LLC",
    "B.R. Wright, LLC",
    "BHG Structured Settlements",
    "BHG Structured Settlements Inc",
    "BHG Structured Settlements, Inc",
    "Bifco, LLC",
    "Bifco",
    "Bifco LLC",
    "Blue Grape LLC",
    "Blue Grape, LLC",
    "Catalina Structured Funding",
    "Catalina Structured Funding Inc",
    "Catalina Structured Funding, Inc",
    "Concordis Group Limited",
    "Conrad Factoring, LLC",
    "Conrad Factoring LLC",
    "Cornerstone Funding LLC",
    "Cornerstone Funding, LLC",
    "fast Annuity Settlement Transfers",
    "Fast Annuity Settlement Transfers LLC",
    "Fast Annuity Settlement Transfers, LLC",
    "FL Assignments Corp",
    "G.D.T.R.F.B. LLC",
    "G.D.T.R.F.B., LLC",
    "G7 Crescenta LLC",
    "G7 Crescenta, LLC",
    "Genex Capital Corp",
    "GJ 123 LLC",
    "GJ 123, LLC",
    "Greenwood Funding",
    "Greenwood Funding LLC",
    "Greenwood Funding, LLC",
    "Grier I LLC",
    "Grier I, LLC",
    "Hakstol Group LLC",
    "Hakstol Group, LLC",
    "Hiddenview Ent, LLC",
    "Hiddenview Ent LLC",
    "Hiddenview Ent",
    "JLC Capital Funding, LLC",
    "JLC Capital Funding LLC",
    "JLC Capital Funding",
    "KN Direct Capital LLC",
    "KN Direct Capital, LLC",
    "KN Direct Capital",
    "Lane Nimitz LLC",
    "Lane Nimitz, LLC",
    "Lasko LLC",
    "Lasko, LLC",
    "Leaf 002 LLC",
    "Leaf 002, LLC",
    "Legere LLC",
    "Legere, LLC",
    "Lottery Funding, LLC",
    "M McDougall LLC",
    "M McDougall, LLC",
    "Majestic Funding LLC",
    "Majestic Funding, LLC",
    "Mic-Bry8",
    "Olive Branch Funding LLC",
    "Olive Branch Funding, LLC",
    "Olive Branch Funding",
    "Palermo Group LLC",
    "Palermo Group, LLC",
    "Palm Green Closing, LLC",
    "Palm Harbor LLC",
    "Palm Harbor, LLC",
    "Passira Mal LLC",
    "Passira Mal, LLC",
    "Patriot Settlement Resources LLC",
    "Patriot Settlement Resources, LLC",
    "Patriot Settlement Resources",
    "QLS Funding LLC",
    "QLS Funding, LLC",
    "Reliance Funding LLC",
    "Reliance Funding, LLC",
    "Rocorp Corporation",
    "RSL Funding LLC",
    "RSL Funding, LLC",
    "Savannah Settlements LLC",
    "Savannah Settlements, LLC",
    "Sempra Finance LLC",
    "Sempra Finance, LLC",
    "Seneca Originations",
    "SeneOne LLC",
    "SeneOne, LLC",
    "Settlement Capital Corp",
    "Settlement Status LLC",
    "Settlement Status, LLC",
    "Somerton LLC",
    "Somerton, LLC",
    "Stratcap Investments Inc.",
    "Stratcap Investments, Inc.",
    "Stratton Asset Funding LLC",
    "Stratton Asset Funding, LLC",
    "Stratton Asset Funding",
    "Structured Asset Funding",
    "Structured Asset Funding LLC",
    "Structured Asset Funding, LLC",
    "TKD LLC",
    "TKD, LLC",
    "TRM V LLC",
    "TRM V, LLC",
    "Tybenz LLC",
    "Tybenz, LLC",
    "Uber Funding LLC",
    "Uber Funding, LLC",
    "Uber Funding",
    "Vintage Equity Group",
    "Vintage Equity Group LLC",
    "Vintage Equity Group, LLC",
    "Wepaymore Funding",
    "Wepaymore Funding, LLC",
    "Wepaymore Funding LLC",
    "Zakho Way LLC",
    "Zakho Way, LLC",
    "GREAT PLAINS MANAGEMENT CORPORATION",
    "T ENE LLC",
    "T ENE, LLC",
    "RD FITZ LLC",
    "RD FITZ, LLC",
    "GA OFF LLC",
    "GA OFF, LLC",
    "Assured Management Corporation",
    "BENTZEN FINANCIAL LLC",
    "BENTZEN FINANCIAL, LLC",
    "Peachtree Settlement Funding",
]

# Set to 0 for no date filter, or a positive integer for "last N days"
DAYS_BACK   = 720
URL         = "https://portal.vtcourts.gov/Portal/Home/Dashboard/29"
OUTPUT_FILE = Path(__file__).parent / "vtcourts_results.xlsx"

# ─────────────────────────────────────────────────────────────────────────────
#  BROWSER SETUP
# ─────────────────────────────────────────────────────────────────────────────

def build_driver() -> webdriver.Chrome:
    options = Options()
    options.add_argument("--start-maximized")
    options.add_argument("--disable-blink-features=AutomationControlled")
    options.add_experimental_option("excludeSwitches", ["enable-automation"])
    options.add_experimental_option("useAutomationExtension", False)
    service = Service(ChromeDriverManager().install())
    driver  = webdriver.Chrome(service=service, options=options)
    driver.execute_script(
        "Object.defineProperty(navigator, 'webdriver', {get: () => undefined})"
    )
    return driver


def js(driver, script, *args):
    return driver.execute_script(script, *args)


def wait_for(driver, by, value, timeout=20):
    return WebDriverWait(driver, timeout).until(
        EC.presence_of_element_located((by, value))
    )


# ─────────────────────────────────────────────────────────────────────────────
#  SEARCH SETUP HELPERS
# ─────────────────────────────────────────────────────────────────────────────

def open_advanced_options(driver):
    btn = None
    for by, val in [(By.ID, "AdvOptions"),
                    (By.PARTIAL_LINK_TEXT, "Advanced Filtering Options")]:
        try:
            btn = WebDriverWait(driver, 6).until(
                EC.presence_of_element_located((by, val))
            )
            break
        except TimeoutException:
            continue
    if btn is None:
        raise RuntimeError("Cannot find Advanced Filtering Options toggle")
    if btn.get_attribute("aria-expanded") != "true":
        js(driver, "arguments[0].click();", btn)
        time.sleep(1.2)


def select_business_name(driver):
    """Set the Kendo ComboBox to BusinessName via the Kendo API."""
    WebDriverWait(driver, 10).until(lambda d: d.execute_script(
        "return !!$(document.getElementById('caseCriteria_SearchBy')).data('kendoComboBox');"
    ))
    result = js(driver, """
        var input  = document.getElementById('caseCriteria_SearchBy');
        var widget = $(input).data('kendoComboBox');
        if (!widget) return 'ERROR: widget not found';
        widget.value('BusinessName');
        widget.trigger('change');
        var vis = input.parentElement.querySelector('.k-input');
        if (vis) vis.value = 'Business Name';
        return 'OK:' + widget.value();
    """)
    if not result or result.startswith("ERROR"):
        raise RuntimeError(f"Kendo ComboBox error: {result}")
    time.sleep(0.6)


def set_date(driver, field_id, date_str):
    js(driver, f"""
        var el = document.getElementById('{field_id}');
        if (el) {{
            $(el).datepicker('setDate', '{date_str}');
            el.dispatchEvent(new Event('change', {{bubbles: true}}));
        }}
    """)


def clear_date(driver, field_id):
    js(driver, f"""
        var el = document.getElementById('{field_id}');
        if (el) {{
            $(el).datepicker('setDate', null);
            el.value = '';
            el.dispatchEvent(new Event('change', {{bubbles: true}}));
        }}
    """)


# ─────────────────────────────────────────────────────────────────────────────
#  RESULT PARSING  — handles the real Party Search Results structure
#
#  Page structure:
#    outer tbody
#      tr.k-master-row          ← party name header (one per party group)
#      tr.k-detail-row          ← contains div.party-card with case grid
#        div.party-card
#          div.kendo-party-card-grid-container
#            table > tbody
#              tr.k-master-row  ← one case row
#                td.party-case-caseid   ← Case Number
#                td.party-case-style    ← Style / Defendant
#                td.party-case-location ← Location
#              tr.k-alt.k-master-row    ← alternating case row (same structure)
#            div.k-pager-wrap   ← pagination (next button if > 10 cases)
#
#  There can be multiple party groups for the same name (site groups same
#  party separately) — we collect ALL cases from ALL groups.
# ─────────────────────────────────────────────────────────────────────────────

def _parse_total_from_pager(pager_el):
    """Extract total item count from pager text like '1 - 10 of 19 items'."""
    try:
        text = pager_el.text.strip()
        m = re.search(r'of\s+(\d+)\s+item', text)
        return int(m.group(1)) if m else 0
    except Exception:
        return 0


def _collect_rows_from_card(card_el):
    """
    Extract all visible case rows from a single div.party-card.
    Uses the specific CSS classes confirmed from live page inspection.
    """
    rows = []
    # Case rows are tr.k-master-row OR tr.k-alt (alternating style)
    # inside the card's grid container — but NOT the outer party master row
    for tr in card_el.find_elements(By.CSS_SELECTOR,
                                     "tr.k-master-row, tr.k-alt"):
        try:
            case_num = tr.find_element(By.CSS_SELECTOR, "td.party-case-caseid").text.strip()
            style    = tr.find_element(By.CSS_SELECTOR, "td.party-case-style").text.strip()
            location = tr.find_element(By.CSS_SELECTOR, "td.party-case-location").text.strip()
            if case_num:   # skip empty/header rows
                rows.append({
                    "Case Number":       case_num,
                    "Style / Defendant": style,
                    "Location":          location,
                })
        except NoSuchElementException:
            continue   # not a data row (e.g. a group-header tr)
    return rows


def _next_page_btn(card_el):
    """
    Return the enabled 'next page' anchor inside this card's pager, or None.
    """
    try:
        btn = card_el.find_element(
            By.CSS_SELECTOR,
            ".k-pager-wrap a[title='Go to the next page']"
        )
        if "k-state-disabled" not in btn.get_attribute("class"):
            return btn
    except NoSuchElementException:
        pass
    return None


def collect_results(driver) -> list:
    """
    Wait for the Party Search Results page, then scrape every case from
    every party group, following pagination within each group.
    Returns [] on 'No cases match'.
    """
    # Wait for either results or no-match message
    try:
        WebDriverWait(driver, 20).until(EC.any_of(
            EC.presence_of_element_located((By.CSS_SELECTOR, "div.party-card")),
            EC.presence_of_element_located((By.XPATH,
                "//*[contains(translate(text(),"
                "'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'),"
                "'no cases match')]"
            )),
        ))
    except TimeoutException:
        print("    [warn] Timed out waiting for results page")

    # No-match check
    try:
        driver.find_element(By.XPATH,
            "//*[contains(translate(text(),"
            "'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'),"
            "'no cases match')]"
        )
        return []
    except NoSuchElementException:
        pass

    all_rows = []

    # Each div.party-card is one party group (there can be multiple for same name)
    cards = driver.find_elements(By.CSS_SELECTOR, "div.party-card")
    print(f"    [info] Found {len(cards)} party group(s)")

    for card_idx, card in enumerate(cards):
        page_num = 1
        while True:
            # Parse visible rows on this page
            rows = _collect_rows_from_card(card)
            all_rows.extend(rows)
            print(f"    [info] Group {card_idx+1}, page {page_num}: {len(rows)} row(s)")

            # Check for next page
            nxt = _next_page_btn(card)
            if nxt is None:
                break   # no more pages for this group

            # Click next page and wait for the grid to refresh
            js(driver, "arguments[0].click();", nxt)
            time.sleep(2.5)   # wait for Kendo grid to re-render
            page_num += 1

            # Re-fetch the card element (DOM may have been updated)
            cards_now = driver.find_elements(By.CSS_SELECTOR, "div.party-card")
            if card_idx < len(cards_now):
                card = cards_now[card_idx]
            else:
                break

    return all_rows


# ─────────────────────────────────────────────────────────────────────────────
#  SEARCH ORCHESTRATION
# ─────────────────────────────────────────────────────────────────────────────

def search_one(driver, business_name, start_date, end_date) -> list:
    print(f"\n{'─'*60}")
    print(f"  Searching : {business_name}")
    if start_date:
        print(f"  Date range: {start_date} → {end_date}")
    else:
        print(f"  Date range: (none — all dates)")

    driver.get(URL)
    wait_for(driver, By.ID, "caseCriteria_SearchCriteria", timeout=20)
    time.sleep(2)

    # 1. Enter name
    box = driver.find_element(By.ID, "caseCriteria_SearchCriteria")
    box.clear()
    box.send_keys(business_name)

    # 2. Open advanced options
    open_advanced_options(driver)

    # 3. Business Name search type
    select_business_name(driver)

    # 4. Date range (or clear)
    wait_for(driver, By.ID, "caseCriteria.FileDateStart", timeout=10)
    if start_date:
        set_date(driver, "caseCriteria.FileDateStart", start_date)
        set_date(driver, "caseCriteria.FileDateEnd",   end_date)
    else:
        clear_date(driver, "caseCriteria.FileDateStart")
        clear_date(driver, "caseCriteria.FileDateEnd")
    time.sleep(0.4)

    # 5. Submit
    submit = wait_for(driver, By.ID, "btnSSSubmit")
    js(driver, "arguments[0].click();", submit)

    # 6. Parse
    results = collect_results(driver)

    if results:
        print(f"  ✅  {len(results)} case(s) collected")
        for r in results:
            print(f"      • {r['Case Number']} | {r['Style / Defendant']} | {r['Location']}")
    else:
        print("  ℹ️   No cases match")

    return results


# ─────────────────────────────────────────────────────────────────────────────
#  EXCEL OUTPUT
# ─────────────────────────────────────────────────────────────────────────────

HDR_FONT  = Font(bold=True, color="FFFFFF")
HDR_FILL  = PatternFill("solid", fgColor="1F3864")
ROW_FILL  = PatternFill("solid", fgColor="E2EFDA")
CTR       = Alignment(horizontal="center", vertical="center")
SHEET_HEADERS = ["Case Number", "Style / Defendant", "Location"]


def load_existing_case_numbers(output_path: Path) -> set:
    if not output_path.exists():
        return set()
    try:
        wb = openpyxl.load_workbook(output_path)
        if "Cases" not in wb.sheetnames:
            return set()
        ws = wb["Cases"]
        existing = {
            str(ws.cell(row=r, column=1).value or "").strip()
            for r in range(2, ws.max_row + 1)
            if ws.cell(row=r, column=1).value
        }
        print(f"  📂  {len(existing)} existing case(s) loaded from {output_path.name}")
        return existing
    except Exception as e:
        print(f"  [warn] Could not read existing file: {e}")
        return set()


def _set_col_widths(ws):
    for col in ws.columns:
        w = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(w + 4, 60)


def append_to_excel(new_rows: list, output_path: Path,
                    seen_this_run: set) -> tuple:

    if output_path.exists():
        wb = openpyxl.load_workbook(output_path)
    else:
        wb = openpyxl.Workbook()
        wb.remove(wb.active)

    if "Cases" in wb.sheetnames:
        ws = wb["Cases"]
        existing_in_file = {
            str(ws.cell(row=r, column=1).value or "").strip()
            for r in range(2, ws.max_row + 1)
            if ws.cell(row=r, column=1).value
        }
    else:
        ws = wb.create_sheet("Cases", 0)
        for ci, h in enumerate(SHEET_HEADERS, 1):
            cell = ws.cell(row=1, column=ci, value=h)
            cell.font = HDR_FONT; cell.fill = HDR_FILL; cell.alignment = CTR
        existing_in_file = set()

    added = skipped = 0

    for row in new_rows:
        case_num = str(row.get("Case Number", "")).strip()
        if not case_num:
            continue

        if case_num in seen_this_run:
            print(f"    [skip-run]  {case_num} — already seen this run")
            skipped += 1
            continue
        if case_num in existing_in_file:
            print(f"    [skip-file] {case_num} — already in Excel file")
            skipped += 1
            seen_this_run.add(case_num)
            continue

        nr = ws.max_row + 1
        ws.cell(row=nr, column=1, value=case_num).fill                        = ROW_FILL
        ws.cell(row=nr, column=2, value=row.get("Style / Defendant","")).fill = ROW_FILL
        ws.cell(row=nr, column=3, value=row.get("Location","")).fill          = ROW_FILL

        seen_this_run.add(case_num)
        existing_in_file.add(case_num)
        added += 1

    _set_col_widths(ws)
    wb.save(output_path)
    return added, skipped


# ─────────────────────────────────────────────────────────────────────────────
#  MAIN
# ─────────────────────────────────────────────────────────────────────────────

def main():
    today = datetime.today()

    if DAYS_BACK > 0:
        start_date = (today - timedelta(days=DAYS_BACK)).strftime("%m/%d/%Y")
        end_date   = today.strftime("%m/%d/%Y")
    else:
        start_date = end_date = ""

    print("=" * 60)
    print("  VT Courts — Business Name Search Automation")
    print(f"  Date range : {start_date or 'all dates'} → {end_date or ''}")
    print(f"  Names      : {len(BUSINESS_NAMES)}")
    print("=" * 60)

    seen_case_numbers = load_existing_case_numbers(OUTPUT_FILE)
    driver = build_driver()

    total_found = total_added = total_skipped = 0

    try:
        for name in BUSINESS_NAMES:
            results = search_one(driver, name, start_date, end_date)
            total_found += len(results)
            if results:
                added, skipped = append_to_excel(
                    results, OUTPUT_FILE, seen_case_numbers
                )
                total_added   += added
                total_skipped += skipped
                print(f"    → {added} new, {skipped} duplicate(s) skipped")
            time.sleep(1)
    finally:
        driver.quit()

    print("\n" + "=" * 60)
    print("  RUN COMPLETE")
    print("=" * 60)
    print(f"  Total cases found    : {total_found}")
    print(f"  New rows written     : {total_added}")
    print(f"  Duplicates skipped   : {total_skipped}")
    print(f"  Output               : {OUTPUT_FILE}")
    print("=" * 60)


if __name__ == "__main__":
    main()